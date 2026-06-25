"""Excel 导入：解析实际数 / 预算数并写入库。

兼容真实业务导出格式：
- 用「日期」列推导年/月（支持 datetime 或 "2026/1/1" 字符串）
- 表头中英文别名（如 "二级产品名称一单一标""签约地一级_汇报""操作目标收入"）
- 脏值（null / 空白 / 未知）归一到「未分配」
- 同维度键重复行按指标求和聚合
- replace=True 时按数据集做覆盖式刷新
另提供 import_workbook：自动识别整本工作簿的多个 sheet 一次导入。
"""

import datetime
import io

from openpyxl import load_workbook
from sqlalchemy import delete
from sqlalchemy.ext.asyncio import AsyncSession

from ..models import Actual, Budget, ImportLog

# 表头别名 -> 标准字段名
HEADER_ALIASES = {
    "year": "year", "年": "year", "年份": "year",
    "month": "month", "月": "month", "月份": "month",
    "date": "date", "日期": "date",
    "region": "region", "洲际": "region", "大区": "region", "区域": "region",
    "签约洲际": "region", "操作洲际": "region",
    "签约地一级_汇报": "region", "操作地一级_汇报": "region",
    "一级汇报": "region", "签约地一级": "region", "操作地一级": "region",
    "l2": "l2", "二级产品": "l2", "二级": "l2", "产品二级": "l2",
    "二级产品名称一单一标": "l2", "二级产品名称": "l2",
    "l3": "l3", "三级产品": "l3", "三级": "l3", "产品三级": "l3",
    "三级产品名称一单一标": "l3", "三级产品名称": "l3",
    "revenue": "revenue", "收入": "revenue", "营收": "revenue",
    "操作目标收入": "revenue", "签约收入": "revenue",
    "gross": "gross", "毛利": "gross", "操作毛利": "gross", "目标毛利": "gross",
    "contrib": "contrib", "贡献利润": "contrib", "贡献": "contrib",
    "签约目标利润": "contrib", "操作目标利润": "contrib",
    "cost": "cost", "成本": "cost", "目标成本": "cost", "操作成本": "cost",
}

# 脏值归一到「未分配」
JUNK_REGIONS = {"", "null", "none", "nan", "(空白)", "（空白）", "未知", "-"}
UNASSIGNED = "未分配"


def _norm(h) -> str | None:
    if h is None:
        return None
    return HEADER_ALIASES.get(str(h).strip(), None)


def _num(v) -> float:
    if v in (None, ""):
        return 0.0
    if isinstance(v, str):
        v = v.replace(",", "").replace("¥", "").strip()
        if v.lower() in ("null", "none", "nan", ""):
            return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _clean_region(v) -> str:
    s = (str(v).strip() if v is not None else "")
    return UNASSIGNED if s.lower() in JUNK_REGIONS else s


def _parse_date(v) -> tuple[int | None, int | None]:
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.year, v.month
    if isinstance(v, str):
        s = v.strip().replace(" ", "")
        for sep in ("/", "-", "."):
            if sep in s:
                parts = s.split(sep)
                try:
                    return int(parts[0]), int(parts[1])
                except (ValueError, IndexError):
                    return None, None
    return None, None


def _rows_from_ws(ws) -> list[dict]:
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return []
    cols = [_norm(h) for h in header]
    out = []
    for r in it:
        rec = {}
        for c, val in zip(cols, r):
            if c:
                rec[c] = val
        out.append(rec)
    return out


def _ym(rec: dict) -> tuple[int | None, int | None]:
    """优先用 date 列推导年月，缺失时回退到 year/month 列。"""
    y, m = _parse_date(rec.get("date"))
    if y is None and rec.get("year"):
        y = int(_num(rec.get("year")))
    if m is None and rec.get("month"):
        m = int(_num(rec.get("month")))
    return y, m


# ---------------- 实际数 ----------------

def _aggregate_actuals(records, default_year):
    agg: dict[tuple, list[float]] = {}
    years: set[int] = set()
    for rec in records:
        y, m = _ym(rec)
        y = y or default_year
        region = _clean_region(rec.get("region"))
        l2 = (str(rec.get("l2")).strip() if rec.get("l2") else "")
        l3 = (str(rec.get("l3")).strip() if rec.get("l3") else "")
        if not (y and m and l2 and l3):
            continue
        years.add(y)
        key = (y, m, region, l2, l3)
        acc = agg.setdefault(key, [0.0, 0.0, 0.0, 0.0])
        acc[0] += _num(rec.get("revenue"))
        acc[1] += _num(rec.get("cost"))
        acc[2] += _num(rec.get("gross"))
        acc[3] += _num(rec.get("contrib"))
    return agg, years


async def _write_actuals(db, agg, years, replace):
    if replace and years:
        await db.execute(delete(Actual).where(Actual.year.in_(years)))
    else:
        for (y, m, region, l2, l3) in agg:
            await db.execute(
                delete(Actual).where(
                    Actual.year == y, Actual.month == m, Actual.region == region,
                    Actual.l2 == l2, Actual.l3 == l3,
                )
            )
    for (y, m, region, l2, l3), (rev, cost, gross, contrib) in agg.items():
        db.add(Actual(year=y, month=m, region=region, l2=l2, l3=l3,
                      revenue=rev, cost=cost, gross=gross, contrib=contrib))
    return len(agg)


async def import_actuals(db, content, filename, default_year=None, replace=False):
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    records = _rows_from_ws(wb.active)
    wb.close()
    agg, years = _aggregate_actuals(records, default_year)
    n = await _write_actuals(db, agg, years, replace)
    db.add(ImportLog(kind="actual", filename=filename, rows=n))
    await db.commit()
    return n


# ---------------- 预算 ----------------

def _aggregate_budgets(records, caliber, default_year):
    # 目标损益表通常无独立成本列，则按 成本=收入−毛利 推导
    has_cost = any("cost" in r for r in records)
    agg: dict[tuple, list[float]] = {}
    for rec in records:
        y, m = _ym(rec)
        y = y or default_year
        region = _clean_region(rec.get("region"))
        l2 = (str(rec.get("l2")).strip() if rec.get("l2") else "")
        if not (y and m and l2):
            continue
        rev = _num(rec.get("revenue"))
        gross = _num(rec.get("gross"))
        cost = _num(rec.get("cost")) if has_cost else (rev - gross)
        key = (y, m, region, l2, caliber)
        acc = agg.setdefault(key, [0.0, 0.0, 0.0, 0.0])
        acc[0] += rev
        acc[1] += cost
        acc[2] += gross
        acc[3] += _num(rec.get("contrib"))
    return agg


async def _write_budgets(db, agg, caliber, year, replace):
    if replace:
        await db.execute(
            delete(Budget).where(Budget.caliber == caliber, Budget.year == year)
        )
    else:
        for (y, m, region, l2, cal) in agg:
            await db.execute(
                delete(Budget).where(
                    Budget.year == y, Budget.month == m, Budget.region == region,
                    Budget.l2 == l2, Budget.caliber == cal,
                )
            )
    for (y, m, region, l2, cal), (rev, cost, gross, contrib) in agg.items():
        db.add(Budget(year=y, month=m, region=region, l2=l2, caliber=cal,
                      revenue=rev, cost=cost, gross=gross, contrib=contrib))
    return len(agg)


async def import_budgets(db, content, filename, caliber, year, replace=False):
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    records = _rows_from_ws(wb.active)
    wb.close()
    agg = _aggregate_budgets(records, caliber, year)
    n = await _write_budgets(db, agg, caliber, year, replace)
    db.add(ImportLog(kind=f"budget-{caliber}", filename=filename, rows=n))
    await db.commit()
    return n


# ---------------- 整本工作簿一键导入 ----------------

def _classify(header_cells) -> tuple[str, str | None]:
    """根据表头识别 sheet 类型：actual / budget-sign / budget-op / unknown。

    预算口径只认专门的目标损益表（签约收入/目标毛利 -> 签约；操作目标收入 -> 操作），
    并跳过旧的简化预算表、含双区域列的辅助表、三级拆分表与备注。
    """
    raw = [str(h).strip() for h in header_cells if h is not None]
    joined = " ".join(raw)
    if any("三级" in h for h in raw):  # 含三级产品 -> 实际数
        return "actual", None
    if "签约收入" in joined or "目标毛利" in joined:  # 专门的签约目标损益
        return "budget", "sign"
    if "操作目标收入" in joined:
        if "签约地" in joined:  # 双区域辅助表，跳过避免与正牌操作目标冲突
            return "unknown", None
        return "budget", "op"
    return "unknown", None


async def import_workbook(db, content, filename, replace=True) -> list[dict]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    results: list[dict] = []
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            continue
        kind, caliber = _classify(header)
        records = _rows_from_ws(ws)
        if kind == "actual":
            agg, years = _aggregate_actuals(records, None)
            n = await _write_actuals(db, agg, years, replace)
            db.add(ImportLog(kind="actual", filename=f"{filename}#{ws.title}", rows=n))
            results.append({"sheet": ws.title, "kind": "actual",
                            "years": sorted(years), "rows": n})
        elif kind == "budget":
            # 预算年份取数据中的众数年（同一 sheet 通常同年）
            yrs = [y for r in records for y, _ in [_ym(r)] if y]
            year = max(set(yrs), key=yrs.count) if yrs else 2026
            agg = _aggregate_budgets(records, caliber, year)
            n = await _write_budgets(db, agg, caliber, year, replace)
            db.add(ImportLog(kind=f"budget-{caliber}",
                             filename=f"{filename}#{ws.title}", rows=n))
            results.append({"sheet": ws.title, "kind": f"budget-{caliber}",
                            "year": year, "rows": n})
        else:
            results.append({"sheet": ws.title, "kind": "skipped", "rows": 0})
    wb.close()
    await db.commit()
    return results
