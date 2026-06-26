"""目标成本（海运先行）：线路主数据 + 时间序列 tracking。

录入以"开一期批量保存"为主力，辅以单条增改与 Excel 导入。
"""

import io
import json
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from pydantic import BaseModel
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from ..db import get_db
from ..models import CostLane, CostTrack

router = APIRouter(prefix="/api/cost", tags=["cost"])
SEED_DIR = Path(__file__).resolve().parents[2] / "seed"


def lane_dict(l: CostLane) -> dict:
    return {
        "id": l.id, "biz_type": l.biz_type, "lane": l.lane,
        "transport_type": l.transport_type,
        "origin_ports": l.origin_ports, "dest_ports": l.dest_ports,
        "warehouse_code": l.warehouse_code, "warehouse_name": l.warehouse_name,
        "warehouse_type": l.warehouse_type, "country": l.country, "region": l.region,
        "pd": l.pd, "carrier": l.carrier, "container_type": l.container_type,
        "unit": l.unit, "currency": l.currency,
        "extra_fee": float(l.extra_fee) if l.extra_fee is not None else None,
        "extra_fee_name": l.extra_fee_name, "note": l.note,
    }


def _scope(q, biz_type, transport_type=None, region=None):
    q = q.where(CostLane.biz_type == biz_type)
    if transport_type:
        q = q.where(CostLane.transport_type == transport_type)
    if region:
        q = q.where(CostLane.region == region)
    return q


@router.get("/lanes")
async def lanes(biz_type: str = "海运", transport_type: str | None = None,
                region: str | None = None, db: AsyncSession = Depends(get_db)):
    rows = (await db.execute(
        _scope(select(CostLane), biz_type, transport_type, region).order_by(CostLane.id)
    )).scalars().all()
    return [lane_dict(l) for l in rows]


class LaneIn(BaseModel):
    biz_type: str = "海运"
    lane: str
    origin_ports: str = ""
    dest_ports: str = ""
    pd: str = ""
    carrier: str = ""
    container_type: str = ""
    unit: str = "FEU"
    currency: str = "USD"
    extra_fee: float | None = None
    extra_fee_name: str = "起运港费用"
    note: str = ""


@router.post("/lanes")
async def create_lane(body: LaneIn, db: AsyncSession = Depends(get_db)):
    obj = CostLane(**body.model_dump())
    db.add(obj)
    await db.commit()
    await db.refresh(obj)
    return lane_dict(obj)


@router.put("/lanes/{lane_id}")
async def update_lane(lane_id: int, body: LaneIn, db: AsyncSession = Depends(get_db)):
    obj = await db.get(CostLane, lane_id)
    if not obj:
        return {"error": "not found"}
    for k, v in body.model_dump().items():
        setattr(obj, k, v)
    await db.commit()
    return lane_dict(obj)


@router.delete("/lanes/{lane_id}")
async def delete_lane(lane_id: int, db: AsyncSession = Depends(get_db)):
    await db.execute(delete(CostTrack).where(CostTrack.lane_id == lane_id))
    obj = await db.get(CostLane, lane_id)
    if obj:
        await db.delete(obj)
    await db.commit()
    return {"ok": True}


@router.get("/trend")
async def trend(biz_type: str = "海运", transport_type: str | None = None,
                region: str | None = None, lane: str | None = None,
                db: AsyncSession = Depends(get_db)):
    """每条线路的目标成本时间序列（喂折线图）。"""
    lq = _scope(select(CostLane), biz_type, transport_type, region)
    if lane:
        lq = lq.where(CostLane.lane == lane)
    lanes = (await db.execute(lq.order_by(CostLane.id))).scalars().all()
    out = []
    for l in lanes:
        ts = (await db.execute(
            select(CostTrack).where(CostTrack.lane_id == l.id)
            .order_by(CostTrack.effective_date)
        )).scalars().all()
        out.append({
            "lane": l.lane, "transport_type": l.transport_type,
            "country": l.country, "unit": l.unit, "currency": l.currency,
            "points": [{
                "effective_date": t.effective_date.isoformat(),
                "end_date": t.end_date.isoformat() if t.end_date else None,
                "amount": float(t.amount), "id": t.id, "source": t.source,
            } for t in ts],
        })
    return {"biz_type": biz_type, "series": out}


@router.get("/current")
async def current(biz_type: str = "海运", transport_type: str | None = None,
                  region: str | None = None, as_of: str | None = None,
                  db: AsyncSession = Depends(get_db)):
    """各线路当前生效目标成本 + 环比上一期涨跌（变动预警）。"""
    today = date.fromisoformat(as_of) if as_of else date.today()
    lanes = (await db.execute(
        _scope(select(CostLane), biz_type, transport_type, region).order_by(CostLane.id)
    )).scalars().all()
    rows = []
    for l in lanes:
        ts = (await db.execute(
            select(CostTrack).where(CostTrack.lane_id == l.id)
            .order_by(CostTrack.effective_date)
        )).scalars().all()
        base = {
            "lane": l.lane, "unit": l.unit, "currency": l.currency,
            "transport_type": l.transport_type, "country": l.country,
            "dest_ports": l.dest_ports, "carrier": l.carrier, "note": l.note,
            "extra_fee": float(l.extra_fee) if l.extra_fee is not None else None,
        }
        if not ts:
            # 无历史价（如空运待录入）：仍列出，标"待录入"
            rows.append({**base, "amount": None, "effective_date": None,
                         "end_date": None, "prev_amount": None, "change_pct": None,
                         "pending": True})
            continue
        cur = None
        for t in ts:
            if t.effective_date <= today and (t.end_date is None or t.end_date >= today):
                cur = t
        if cur is None:
            cur = ts[-1]  # 兜底取最新一期
        idx = ts.index(cur)
        prev = ts[idx - 1] if idx > 0 else None
        chg = None
        if prev and float(prev.amount):
            chg = round((float(cur.amount) - float(prev.amount)) / float(prev.amount) * 100, 1)
        rows.append({
            **base,
            "amount": float(cur.amount),
            "effective_date": cur.effective_date.isoformat(),
            "end_date": cur.end_date.isoformat() if cur.end_date else None,
            "prev_amount": float(prev.amount) if prev else None,
            "change_pct": chg, "pending": False,
        })
    return {"biz_type": biz_type, "as_of": today.isoformat(), "rows": rows}


@router.get("/last_period")
async def last_period(biz_type: str = "海运", transport_type: str | None = None,
                      region: str | None = None, db: AsyncSession = Depends(get_db)):
    """开新一期时预填：各线路最新一期的价格 + 建议的新生效日期。"""
    lanes = (await db.execute(
        _scope(select(CostLane), biz_type, transport_type, region).order_by(CostLane.id)
    )).scalars().all()
    rows = []
    latest_eff = None   # 最近生效的一期(用它的结束日推下一期, 避免被个别长效期带偏)
    latest_eff_end = None
    for l in lanes:
        t = (await db.execute(
            select(CostTrack).where(CostTrack.lane_id == l.id)
            .order_by(CostTrack.effective_date.desc()).limit(1)
        )).scalars().first()
        rows.append({
            "lane_id": l.id, "lane": l.lane, "unit": l.unit, "currency": l.currency,
            "transport_type": l.transport_type, "country": l.country,
            "region": l.region, "note": l.note,
            "last_amount": float(t.amount) if t else None,
            "last_end": t.end_date.isoformat() if t and t.end_date else None,
        })
        if t and (latest_eff is None or t.effective_date > latest_eff):
            latest_eff = t.effective_date
            latest_eff_end = t.end_date
    base = latest_eff_end or latest_eff
    suggest_eff = (base + timedelta(days=1)).isoformat() if base else None
    return {"biz_type": biz_type, "rows": rows, "suggest_effective": suggest_eff}


class TrackItem(BaseModel):
    lane_id: int
    amount: float | None = None


class BatchIn(BaseModel):
    biz_type: str = "海运"
    effective_date: str
    end_date: str | None = None
    items: list[TrackItem]


@router.post("/track/batch")
async def save_batch(body: BatchIn, db: AsyncSession = Depends(get_db)):
    """开一期批量保存：跳过空值，按 (线路,生效日期) 覆盖。"""
    eff = date.fromisoformat(body.effective_date)
    end = date.fromisoformat(body.end_date) if body.end_date else None
    n = 0
    for it in body.items:
        if it.amount is None:
            continue
        # 同线路同生效日期 → 覆盖
        existing = (await db.execute(
            select(CostTrack).where(
                CostTrack.lane_id == it.lane_id, CostTrack.effective_date == eff
            )
        )).scalars().first()
        if existing:
            existing.amount = it.amount
            existing.end_date = end
            existing.source = "手工"
        else:
            db.add(CostTrack(lane_id=it.lane_id, effective_date=eff, end_date=end,
                             amount=it.amount, source="手工"))
        n += 1
    await db.commit()
    return {"saved": n}


class TrackIn(BaseModel):
    lane_id: int
    effective_date: str
    end_date: str | None = None
    amount: float


@router.put("/track/{tid}")
async def update_track(tid: int, body: TrackIn, db: AsyncSession = Depends(get_db)):
    obj = await db.get(CostTrack, tid)
    if not obj:
        return {"error": "not found"}
    obj.lane_id = body.lane_id
    obj.effective_date = date.fromisoformat(body.effective_date)
    obj.end_date = date.fromisoformat(body.end_date) if body.end_date else None
    obj.amount = body.amount
    obj.source = "手工"
    await db.commit()
    return {"ok": True}


@router.delete("/track/{tid}")
async def delete_track(tid: int, db: AsyncSession = Depends(get_db)):
    obj = await db.get(CostTrack, tid)
    if obj:
        await db.delete(obj)
        await db.commit()
    return {"ok": True}


# ---------- 导出: 还原成与原始 tracking 表一致的 3 sheet xlsx ----------
async def _lane_tracks(db, biz_type):
    """返回 {lane_id: lane}, {lane_id: [tracks asc]} (限定 biz_type)。"""
    lanes = (await db.execute(
        select(CostLane).where(CostLane.biz_type == biz_type)
    )).scalars().all()
    lmap = {l.id: l for l in lanes}
    tmap = defaultdict(list)
    if lmap:
        ts = (await db.execute(
            select(CostTrack).where(CostTrack.lane_id.in_(list(lmap)))
            .order_by(CostTrack.effective_date)
        )).scalars().all()
        for t in ts:
            tmap[t.lane_id].append(t)
    return lmap, tmap


@router.get("/export")
async def export(db: AsyncSession = Depends(get_db)):
    wb = Workbook()
    wb.remove(wb.active)

    # ===== 跨境到仓 (逐仓逐期; 区域价铺到每个仓) =====
    cb_lanes, cb_tracks = await _lane_tracks(db, "跨境到仓")
    # (region, transport) -> tracks
    rt = {}
    for lid, l in cb_lanes.items():
        rt[(l.region, l.transport_type)] = cb_tracks.get(lid, [])
    ws = wb.create_sheet("跨境到仓")
    ws.append(["生效日期", "结束日期", "起运仓", "起运港（三字码）", "目的港（三字码）",
               "目的仓名称", "目的仓编码", "目的地仓类型（京东仓/非京东仓）", "运输类型（空运/海运）",
               "PD", "目标成本单价A", "计费单位", "币种", "目标成本单价B", "计费单位", "币种"])
    cb_detail = json.loads((SEED_DIR / "cb_detail.json").read_text(encoding="utf-8"))
    for w in cb_detail:
        for t in rt.get((w["region"], w["transport_type"]), []):
            ws.append([t.effective_date, t.end_date, "", w["origin_port"], w["dest_port"],
                       w["wh_name"], w["wh_code"], w["wh_type"], w["transport_type"],
                       w["pd"], float(t.amount), w["unit"], w["currency"], "", "", ""])

    # ===== 海运 (线路级, 直接对应原表) =====
    hy_lanes, hy_tracks = await _lane_tracks(db, "海运")
    ws = wb.create_sheet("海运")
    ws.append(["生效日期", "结束日期", "线路", "起运港（英文逗号）", "目的港（英文逗号）",
               "PD", "船东（英文逗号）", "柜型", "目标成本", "计费单位", "币种", "起运港费用"])
    for lid, l in hy_lanes.items():
        for t in hy_tracks.get(lid, []):
            ws.append([t.effective_date, t.end_date, l.lane, l.origin_ports, l.dest_ports,
                       l.pd, l.carrier, l.container_type, float(t.amount), l.unit, l.currency,
                       float(l.extra_fee) if l.extra_fee is not None else None])

    # ===== 小包 (逐PD逐期; 线路 all-in 铺到每个PD, 揽收/库内/清关取固定) =====
    sp_lanes, sp_tracks = await _lane_tracks(db, "小包")
    line_tracks = {}
    for lid, l in sp_lanes.items():
        line_tracks[l.region] = sp_tracks.get(lid, [])
    ws = wb.create_sheet("小包")
    ws.append(["生效日期", "结束日期", "线路", "PD", "目的国（二字码）",
               "A-揽收", "计费单位", "币种", "A-库内", "计费单位", "币种",
               "B-all in", "计费单位", "币种", "C-清关", "计费单位", "币种"])
    sp_detail = json.loads((SEED_DIR / "sp_detail.json").read_text(encoding="utf-8"))
    for p in sp_detail:
        for t in line_tracks.get(p["region"], []):
            ws.append([t.effective_date, t.end_date, p["region"], p["pd"], p["country"],
                       p["lanshou"], p["lanshou_unit"], p["lanshou_cur"],
                       p["kunei"], p["kunei_unit"], p["kunei_cur"],
                       float(t.amount), p["allin_unit"], p["allin_cur"],
                       p["qingguan"], p["qingguan_unit"], p["qingguan_cur"]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=target_cost_tracking.xlsx"},
    )
